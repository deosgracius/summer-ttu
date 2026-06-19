"""Parse admin-provided registrar Excel files into campus data.

Two sheet shapes are recognized (see the real sample,
`CMPE SUM26_F26 DEPARTMENTAL ELECTIVES LIST.xlsx`):

* **Schedule sheets** — one header row starting with `CRN`, then one row per
  offered section. Note/disclaimer rows are mixed in and must be skipped.
* **Master list** — a wide reference matrix of electives by category.

Everything is best-effort and defensive: a malformed row is skipped and reported
rather than aborting the whole import.
"""
import io
import re
from openpyxl import load_workbook

# Canonical schedule columns, in order.
SCHEDULE_COLS = [
    "crn", "subject", "course", "section", "title", "permit_required",
    "days", "times", "start_date", "end_date", "campus", "building",
    "room_number", "instructor", "max_enroll",
]


def _s(v):
    return "" if v is None else str(v).strip()


def _split_prereqs(title: str):
    """'Network Analysis (ECE 3302)' -> ('Network Analysis', 'ECE 3302')."""
    title = title.strip()
    start = title.find("(")
    end = title.rfind(")")
    if start != -1 and end > start:
        clean = title[:start].strip()
        prereqs = title[start + 1:end].strip()
        return clean, prereqs
    return title, ""


def _is_greenish(cell) -> bool:
    """Best-effort: graduate courses are highlighted green. Detect a solid
    fill whose RGB is clearly green-dominant. Falls back to False on anything
    unexpected (theme colors, no fill, etc.)."""
    try:
        fill = cell.fill
        if not fill or fill.patternType != "solid":
            return False
        rgb = getattr(fill.fgColor, "rgb", None)
        if not isinstance(rgb, str) or len(rgb) < 6:
            return False
        hexpart = rgb[-6:]
        r, g, b = int(hexpart[0:2], 16), int(hexpart[2:4], 16), int(hexpart[4:6], 16)
        return g > 140 and g > r + 30 and g > b + 30
    except Exception:
        return False


def _find_header_row(ws, max_scan=8):
    """Return the 1-based row index whose first cell is 'CRN', or None."""
    for r in range(1, min(max_scan, ws.max_row) + 1):
        if _s(ws.cell(row=r, column=1).value).upper() == "CRN":
            return r
    return None


def _parse_schedule(ws, semester: str):
    rows, skipped = [], []
    header = _find_header_row(ws)
    if not header:
        return rows, skipped
    for r in range(header + 1, ws.max_row + 1):
        vals = [ws.cell(row=r, column=c).value for c in range(1, 16)]
        cells = [_s(v) for v in vals]
        if not any(cells):
            continue
        # A real section row has a subject and a course number. Note/disclaimer
        # rows (e.g. '***ECE COURSES HIGHLIGHTED...') fail this and get skipped.
        subject, course = cells[1], cells[2]
        if not subject or not course:
            joined = " ".join(c for c in cells if c)
            skipped.append(joined[:120])
            continue
        title_clean, prereqs = _split_prereqs(cells[4])
        try:
            max_enroll = int(float(cells[14])) if cells[14] else 0
        except ValueError:
            max_enroll = 0
        rows.append({
            "crn": cells[0], "subject": subject, "course": course,
            "section": cells[3], "title": title_clean, "prerequisites": prereqs,
            "permit_required": cells[5], "days": cells[6], "times": cells[7],
            "start_date": cells[8], "end_date": cells[9], "campus": cells[10],
            "building": cells[11], "room_number": cells[12], "instructor": cells[13],
            "max_enroll": max_enroll,
            "is_graduate": _is_greenish(ws.cell(row=r, column=2)),
            "semester": semester,
        })
    return rows, skipped


# Column triples in the master list: (category, code_col, title_col).
_MASTER_TRIPLES = [("ECE", 2, 3), ("CS/MATH", 4, 5), ("Group B Project Lab", 6, 7)]


def _parse_master(ws):
    rows = []
    year = ""
    for r in range(1, min(6, ws.max_row) + 1):
        for c in range(1, ws.max_column + 1):
            v = _s(ws.cell(row=r, column=c).value)
            if re.match(r"^\d{4}\s*[-–]\s*\d{4}$", v):
                year = v.replace(" ", "")
    for r in range(2, ws.max_row + 1):
        for category, ccol, tcol in _MASTER_TRIPLES:
            code = _s(ws.cell(row=r, column=ccol).value)
            title_raw = _s(ws.cell(row=r, column=tcol).value)
            # A catalog entry is a course code like 'ECE 3306' / 'CS 3352*'.
            if not re.match(r"^[A-Z]{2,4}\s*\d{4}\*?$", code):
                continue
            title_clean, prereqs = _split_prereqs(title_raw)
            rows.append({
                "category": category, "code": code.replace("*", "").strip(),
                "title": title_clean, "prerequisites": prereqs,
                "notes": "honors/restricted" if "*" in code else "",
                "catalog_year": year,
            })
    return rows


def parse_workbook(file_bytes: bytes, semester_override: str = ""):
    """Parse an uploaded .xlsx into a preview dict (no DB writes).

    Returns: {sheets: [{name, type, count, skipped}], offerings: [...],
              catalog: [...]}. `semester` per offering defaults to the sheet
              name unless `semester_override` is given.
    """
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    offerings, catalog, sheet_info = [], [], []
    for ws in wb.worksheets:
        name = ws.title
        if "MASTER" in name.upper():
            rows = _parse_master(ws)
            catalog.extend(rows)
            sheet_info.append({"name": name, "type": "catalog",
                               "count": len(rows), "skipped": 0})
        elif _find_header_row(ws):
            sem = semester_override or name
            rows, skipped = _parse_schedule(ws, sem)
            offerings.extend(rows)
            sheet_info.append({"name": name, "type": "offerings",
                               "count": len(rows), "skipped": len(skipped),
                               "skipped_samples": skipped[:5]})
        else:
            sheet_info.append({"name": name, "type": "unrecognized",
                               "count": 0, "skipped": 0})
    return {"sheets": sheet_info, "offerings": offerings, "catalog": catalog}
